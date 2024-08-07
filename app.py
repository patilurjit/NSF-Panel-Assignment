import streamlit as st
import numpy as np
import pulp
import pandas as pd
import time
from io import BytesIO
import base64
import openpyxl

def review_opt():
    prob = pulp.LpProblem("ReviewAssignment", pulp.LpMinimize)

    x = pulp.LpVariable.dicts('x', range(num_vars), lowBound = 0, upBound = 1, cat = 'Binary')

    prob += pulp.lpSum([f[i] * x[i] for i in range(num_vars)])

    for j in range(num_proposals):
        prob += pulp.lpSum([x[i*num_proposals + j] for i in range(num_reviewers)]) == reviews_per_proposal

    for i in range(num_reviewers):
        prob += pulp.lpSum([x[i*num_proposals + j] for j in range(num_proposals)]) <= max_reviews_per_reviewer
        prob += pulp.lpSum([x[i*num_proposals + j] for j in range(num_proposals)]) >= min_reviews_per_reviewer

    for i in range(num_reviewers):
        for j in range(num_proposals):
            if rankings[i, j] == 0:
                prob += x[i*num_proposals + j] == 0

    prob.solve()

    if pulp.LpStatus[prob.status] != 'Optimal':
        st.error('No optimal solution found. Please try again with different inputs.')
        return None, None
    
    assignments = np.zeros((num_reviewers, num_proposals))
    for i in range(num_reviewers):
        for j in range(num_proposals):
            assignments[i, j] = pulp.value(x[i*num_proposals + j])

    fval = prob.objective.value()

    return assignments, fval

def lead_opt(lead_counts, assignments):
    prob = pulp.LpProblem("LeadAssignment", pulp.LpMinimize)

    x = pulp.LpVariable.dicts('x', range(num_vars), lowBound=0, upBound=1, cat='Binary')

    prob += pulp.lpSum([f[i] * x[i] for i in range(num_vars)])

    for i in range(num_reviewers):
        prob += pulp.lpSum([x[i*num_proposals + j] for j in range(num_proposals)]) <= lead_counts.flatten()[i]

    for i in range(num_reviewers):
        prob += pulp.lpSum([x[i*num_proposals + j] for j in range(num_proposals)]) >= 1

    for i in range(num_vars):
        prob += x[i] <= assignments.flatten()[i]

    for j in range(num_proposals):
        prob += pulp.lpSum([x[i*num_proposals + j] for i in range(num_reviewers)]) == 1

    prob.solve()

    if pulp.LpStatus[prob.status] != 'Optimal':
        st.error('No optimal solution found. Please try again with different inputs.')
        return None, None
    
    lead_assignments = np.zeros((num_reviewers, num_proposals))
    for i in range(num_reviewers):
        for j in range(num_proposals):
            lead_assignments[i, j] = pulp.value(x[i*num_proposals + j])

    lead_assignments = lead_assignments.T

    lead_fval = prob.objective.value()

    return lead_assignments, lead_fval

def scribe_opt(max_scribe, min_scribe, assignments):
    prob = pulp.LpProblem("ScribeAssignment", pulp.LpMinimize)

    x = pulp.LpVariable.dicts('x', range(num_vars), lowBound=0, upBound=1, cat='Binary')

    prob += pulp.lpSum([f[i] * x[i] for i in range(num_vars)])

    for i in range(num_reviewers):
        prob += pulp.lpSum(x[i*num_proposals + j] for j in range(num_proposals)) <= max_scribe
        prob += pulp.lpSum(x[i*num_proposals + j] for j in range(num_proposals)) >= min_scribe

    for j in range(num_proposals):
        prob += pulp.lpSum(x[i*num_proposals + j] for i in range(num_reviewers)) == 1

    for i in range(num_reviewers):
        for j in range(num_proposals):
            if assignments[i, j] == 0:
                prob += x[i*num_proposals + j] == 0

    prob.solve()

    if pulp.LpStatus[prob.status] != 'Optimal':
        st.error('No optimal solution found. Please try again with different inputs.')
        return None, None
    
    scribe_assignments = np.zeros((num_reviewers, num_proposals))
    for i in range(num_reviewers):
        for j in range(num_proposals):
            scribe_assignments[i, j] = pulp.value(x[i*num_proposals + j])

    scribe_assignments = scribe_assignments.T
    
    scribe_fval = prob.objective.value()

    return scribe_assignments, scribe_fval

def check_review_conflicts(assignments):
    conflicts = (assignments.astype(bool) & (rankings == 0))
    if np.any(conflicts):
        st.error('Conflicts detected in the assignments:')
        reviewer_idx, proposal_idx = np.where(conflicts)
        for k in range(len(reviewer_idx)):
            st.error(f'Reviewer {reviewer_idx[k+1]} assigned to Proposal {proposal_idx[k+1]} (conflict) \n')
    else:
        st.sidebar.success('No conflicts found in the review assignments.')

def check_lead_conflicts(assignments):
    conflicts = (assignments.T.astype(bool) & (rankings == 0))
    if np.any(conflicts):
        st.error('Conflicts detected in the assignments:')
        reviewer_idx, proposal_idx = np.where(conflicts)
        for k in range(len(reviewer_idx)):
            st.error(f'Reviewer {reviewer_idx[k+1]} assigned to Proposal {proposal_idx[k+1]} (conflict) \n')
    else:
        st.sidebar.success('No conflicts found in the lead assignments.')

def check_scribe_conflicts(assignments):
    conflicts = (assignments.T.astype(bool) & (rankings == 0))
    if np.any(conflicts):
        st.error('Conflicts detected in the assignments:')
        reviewer_idx, proposal_idx = np.where(conflicts)
        for k in range(len(reviewer_idx)):
            st.error(f'Reviewer {reviewer_idx[k+1]} assigned to Proposal {proposal_idx[k+1]} (conflict) \n')
    else:
        st.sidebar.success('No conflicts found in the scribe assignments.')

def check_review_count(assignments):
    proposal_count = np.sum(assignments, 0)

    for proposal in range(num_proposals):
        if proposal_count[proposal] != reviews_per_proposal:
            st.error(f'Proposal {proposal+1} does not have the required number of reviews.')
        
def check_lead_count(assignments):
    proposal_count = np.sum(assignments, 0)

    for proposal in range(num_proposals):
        if proposal_count[proposal] != 1:
            st.error(f'Proposal {proposal+1} does not have the required number of leads.')

def check_scribe_count(assignments):
    proposal_count = np.sum(assignments, 0)

    for proposal in range(num_proposals):
        if proposal_count[proposal] != 1:
            st.error(f'Proposal {proposal+1} does not have the required number of scribes.')

def make_combined_assignments(assignments, l_assignments, rankings):
    combined_assignments = np.full((num_proposals, num_reviewers), '-', dtype = object)
    fval_assignments = np.zeros((num_proposals, num_reviewers))

    for proposal in range(num_proposals):
        for reviewer in range(num_reviewers):
            if l_assignments[proposal, reviewer] == 1:
                combined_assignments[proposal, reviewer] = 'LSR'
                fval_assignments[proposal, reviewer] = 2+3+1
            elif assignments[reviewer, proposal] == 1:
                combined_assignments[proposal, reviewer] = 'R'
                fval_assignments[proposal, reviewer] = 1
            elif rankings[reviewer, proposal] == 0:
                combined_assignments[proposal, reviewer] = 'COI'
            else:
                combined_assignments[proposal, reviewer] = '-'

    fval_combined = np.sum(rankings.T * fval_assignments)

    return combined_assignments, fval_combined

def make_combined_assignments_with_scribe(assignments, l_assignments, s_assignments, rankings):
    new_assignments1 = assignments.T - l_assignments
    combined_assignments = np.full((num_proposals, num_reviewers), '-', dtype = object)
    fval_assignments = np.zeros((num_proposals, num_reviewers))

    for proposal in range(num_proposals):
        for reviewer in range(num_reviewers):
            if s_assignments[proposal, reviewer] == 1:
                combined_assignments[proposal, reviewer] = 'Scribe'
                fval_assignments[proposal, reviewer] = 2+1
            elif l_assignments[proposal, reviewer] == 1:
                combined_assignments[proposal, reviewer] = 'Lead'
                fval_assignments[proposal, reviewer] = 3+1
            elif new_assignments1[proposal, reviewer] == 1:
                combined_assignments[proposal, reviewer] = 'R'
                fval_assignments[proposal, reviewer] = 1
            elif rankings[reviewer, proposal] == 0:
                combined_assignments[proposal, reviewer] = 'Conflict'
            else:
                combined_assignments[proposal, reviewer] = '-'

    fval_combined = np.sum(rankings.T * fval_assignments)

    return combined_assignments, fval_combined

def calculate_fairness_metrics(assignments, l_assignments, rankings, lead_counts):
    proposal_count = np.sum(assignments, 0)
    fairness_prop_metric = np.zeros(num_proposals)

    for proposal in range(num_proposals):
        fairness_prop_metric[proposal] = np.sum(assignments[:, proposal]*rankings[:, proposal])/proposal_count[proposal]

    reviews_count = np.sum(assignments, 1)
    fairness_metric = np.zeros(num_reviewers)

    for reviewer in range(num_reviewers):
        fairness_metric[reviewer] = np.sum(assignments[reviewer, :] * rankings[reviewer, :]) / reviews_count[reviewer]

    fairness_lsr_metric = np.zeros(num_reviewers)

    for reviewer in range(num_reviewers):
        fairness_lsr_metric[reviewer] = np.sum(l_assignments[:, reviewer] * rankings[reviewer, :]) / lead_counts[reviewer]

    return fairness_prop_metric, fairness_metric, fairness_lsr_metric

def calculate_fairness_metrics_with_scribe(assignments, l_assignments, s_assignments, rankings, lead_counts, scribe_counts):
    proposal_count = np.sum(assignments, 0)
    fairness_prop_metric = np.zeros(num_proposals)

    for proposal in range(num_proposals):
        fairness_prop_metric[proposal] = np.sum(assignments[:, proposal]*rankings[:, proposal])/proposal_count[proposal]

    reviews_count = np.sum(assignments, 1)
    fairness_metric = np.zeros(num_reviewers)

    for reviewer in range(num_reviewers):
        fairness_metric[reviewer] = np.sum(assignments[reviewer, :] * rankings[reviewer, :]) / reviews_count[reviewer]

    fairness_l_metric = np.zeros(num_reviewers)

    for reviewer in range(num_reviewers):
        fairness_l_metric[reviewer] = np.sum(l_assignments[:, reviewer] * rankings[reviewer, :]) / lead_counts[reviewer]

    fairness_s_metric = np.zeros(num_reviewers)

    for reviewer in range(num_reviewers):
        fairness_s_metric[reviewer] = np.sum(s_assignments[:, reviewer] * rankings[reviewer, :]) / scribe_counts[reviewer]

    return fairness_prop_metric, fairness_metric, fairness_l_metric, fairness_s_metric

def generate_conflict_reordered_assignments(assignments, l_assignments, rankings):
    proposal_indices = list(range(num_proposals))

    con_reordered_combined_assignments = np.full((num_proposals, num_reviewers), '-', dtype = object)
    reviewer_conflict_matrix = (rankings == 0)

    grouped_proposals = []
    for reviewer in range(num_reviewers):
        conflicts = np.where(reviewer_conflict_matrix[reviewer, :])[0].tolist()
        grouped_proposals.extend(conflicts)

    grouped_proposals = list(dict.fromkeys(grouped_proposals))
    remaining_proposals = [item for item in proposal_indices if item not in grouped_proposals]
    grouped_proposals.extend(remaining_proposals)

    for i in range(num_proposals):
        proposal = grouped_proposals[i]
        for reviewer in range(num_reviewers):
            if l_assignments[proposal, reviewer] == 1:
                con_reordered_combined_assignments[i, reviewer] = 'LSR'
            elif assignments[reviewer, proposal] == 1:
                con_reordered_combined_assignments[i, reviewer] = 'R'
            elif rankings[reviewer, proposal] == 0:
                con_reordered_combined_assignments[i, reviewer] = 'COI'
            else:
                con_reordered_combined_assignments[i, reviewer] = '-'

    column_names = [f'Reviewer {i}' for i in range(1, num_reviewers + 1)]
    row_names = [f'Proposal {i+1}' for i in grouped_proposals]
    con_reordered_combined_assignments_df = pd.DataFrame(con_reordered_combined_assignments, columns=column_names, index=row_names)

    return con_reordered_combined_assignments_df

def generate_conflict_reordered_assignments_with_scribe(assignments, l_assignments, s_assignments, rankings):
    proposal_indices = list(range(num_proposals))

    con_reordered_combined_assignments_with_scribe = np.full((num_proposals, num_reviewers), '-', dtype = object)
    reviewer_conflict_matrix = (rankings == 0)

    grouped_proposals = []
    for reviewer in range(num_reviewers):
        conflicts = np.where(reviewer_conflict_matrix[reviewer, :])[0].tolist()
        grouped_proposals.extend(conflicts)

    grouped_proposals = list(dict.fromkeys(grouped_proposals))
    remaining_proposals = [item for item in proposal_indices if item not in grouped_proposals]
    grouped_proposals.extend(remaining_proposals)

    for i in range(num_proposals):
        proposal = grouped_proposals[i]
        for reviewer in range(num_reviewers):
            if l_assignments[proposal, reviewer] == 1:
                con_reordered_combined_assignments_with_scribe[i, reviewer] = 'L'
            elif s_assignments[proposal, reviewer] == 1:
                con_reordered_combined_assignments_with_scribe[i, reviewer] = 'S'
            elif assignments[reviewer, proposal] == 1:
                con_reordered_combined_assignments_with_scribe[i, reviewer] = 'R'
            elif rankings[reviewer, proposal] == 0:
                con_reordered_combined_assignments_with_scribe[i, reviewer] = 'COI'
            else:
                con_reordered_combined_assignments_with_scribe[i, reviewer] = '-'

    column_names = [f'Reviewer {i}' for i in range(1, num_reviewers + 1)]
    row_names = [f'Proposal {i+1}' for i in grouped_proposals]
    con_reordered_combined_assignments_df = pd.DataFrame(con_reordered_combined_assignments_with_scribe, columns=column_names, index=row_names)

    return con_reordered_combined_assignments_df

def generate_leads_reordered_assignments(l_assignments, c_assignments):
    num_proposals = l_assignments.shape[0]
    num_reviewers = l_assignments.shape[1]

    total_leads = np.sum(l_assignments, 0)

    proposal_discussion_order = []

    discussed_proposals = np.zeros(num_proposals, dtype = bool)

    sorted_reviewer_indices = np.argsort(-total_leads)

    for reviewer_index in range(num_reviewers):
        reviewer = sorted_reviewer_indices[reviewer_index]
        reviewer_leads = np.where(l_assignments[:, reviewer] == 1)[0]
        for proposal in reviewer_leads:
            if not discussed_proposals[proposal]:
                proposal_discussion_order.append(proposal)
                discussed_proposals[proposal] = True

    sorted_combined_assignments = np.full((num_proposals, num_reviewers), '-', dtype = object)

    for i in range(num_proposals):
        original_proposal = proposal_discussion_order[i]
        sorted_combined_assignments[i, :] = c_assignments[original_proposal, :]

    column_names = [f'Reviewer {i}' for i in range(1, num_reviewers + 1)]
    row_names = [f'Proposal {i+1}' for i in proposal_discussion_order]
    lead_reordered_combined_assignments_df = pd.DataFrame(sorted_combined_assignments, columns = column_names, index = row_names)

    return lead_reordered_combined_assignments_df

def generate_rating_reordered_assignments(c_assignments, assignments):
    column_names = [f'Reviewer {i}' for i in range(1, num_reviewers + 1)]
    row_names = [f'Proposal {i+1}' for i in range(num_proposals)]
    
    if 'ratings' in st.session_state:
        ratings_table = st.session_state['ratings']
    
    else:
        ratings = ['E', 'V', 'G', 'F', 'P', 'E/V', 'V/G', 'G/F', 'F/P']

        ratings_matrix = np.full((num_proposals, num_reviewers), '-', dtype = object)

        for proposal in range(num_proposals):
            for reviewer in range(num_reviewers):
                if assignments[reviewer, proposal] != 0:
                    random_rating = np.random.choice(ratings)
                    ratings_matrix[proposal, reviewer] = random_rating
                else:
                    ratings_matrix[proposal, reviewer] = '-'

        ratings_table = pd.DataFrame(ratings_matrix, columns = column_names, index = row_names)

    ratings_matrix = ratings_table.to_numpy(dtype = str)

    rating_to_points = {
            'E': 1, 'V': 2, 'G': 3, 'F': 4, 'P': 5,
            'E/V': (1 + 2) / 2, 'V/G': (2 + 3) / 2, 'G/F': (3 + 4) / 2, 'F/P': (4 + 5) / 2
        }

    points_matrix = np.empty((num_proposals, num_reviewers), dtype = object)

    for proposal in range(num_proposals):
            for reviewer in range(num_reviewers):
                rating = ratings_matrix[proposal, reviewer]
                if rating:
                    points_matrix[proposal, reviewer] = str(rating_to_points.get(rating, ''))
                else:
                    points_matrix[proposal, reviewer] = ''

    numeric_points_matrix = np.full((num_proposals, num_reviewers), np.nan)
    for proposal in range(num_proposals):
        for reviewer in range(num_reviewers):
            if points_matrix[proposal, reviewer]:
                numeric_points_matrix[proposal, reviewer] = float(points_matrix[proposal, reviewer])

    total_points_per_proposal = np.nansum(numeric_points_matrix, axis = 1)
    sorted_indices = np.argsort(total_points_per_proposal)
    sorted_combined_assignments = c_assignments[sorted_indices, :]
    total_points_for_sorted_proposals = total_points_per_proposal[sorted_indices]

    new_row_names = [f'Proposal {i+1}' for i in sorted_indices]

    sorted_combined_assignments_df = pd.DataFrame(sorted_combined_assignments, columns=column_names, index = new_row_names)
    total_points_df = pd.DataFrame(total_points_for_sorted_proposals, columns = ['Total Points'], index = new_row_names)

    rating_combined_assignments_with_total = pd.concat([sorted_combined_assignments_df, total_points_df], axis = 1)

    return ratings_table, rating_combined_assignments_with_total

def generate_rankings_matrix_template(num_reviewers, num_proposals):
    rankings = np.zeros((num_proposals, num_reviewers))
    column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
    row_names = [f'Proposal {i+1}' for i in range(num_proposals)]
    rankings_sheet = pd.DataFrame(rankings, columns = column_names, index = row_names)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine = 'openpyxl') as writer:
        rankings_sheet.to_excel(writer, index = True, sheet_name = 'Rankings Matrix')

    buffer.seek(0)
    return buffer

def generate_ratings_matrix_template(num_reviewers, num_proposals):
    reference = np.array([['E', 'E/V', '-', 'V', 'V/G', 'G/F'],
                        ['V', 'V/G', 'P', 'F/P', '-', 'F'],
                        ['G', 'E/V', 'F/P', '-', 'G/F', 'F'],
                        ['-', 'V', 'E/V', 'V', 'V/G', 'V'],
                        ['F', 'G/F', 'G/F', 'F', '-', 'E/V'],
                        ['-', 'G/F', 'F/P', 'G', 'F/P', 'P'],
                        ['E', 'F', 'P', 'G', 'V', '-'],
                        ['F', 'G/F', 'F/P', '-', 'F/P', 'V'],
                        ['G/F', 'E/V', '-', 'E', 'G', 'E'],
                        ['G', 'E/V', 'F', 'E/V', 'G', '-'],
                        ['G', '-', 'P', 'F/P', 'E', 'F/P'],
                        ['V', '-', 'G/F', 'G', 'E', 'G']], dtype=object)
    reference_df = pd.DataFrame(reference, columns = [f'Reviewer {i+1}' for i in range(reference.shape[1])], index = [f'Proposal {i+1}' for i in range(reference.shape[0])])
    
    ratings = np.zeros((num_proposals, num_reviewers))
    column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
    row_names = [f'Proposal {i+1}' for i in range(num_proposals)]
    ratings_sheet = pd.DataFrame(ratings, columns = column_names, index = row_names)

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine = 'openpyxl') as writer:
        reference_df.to_excel(writer, index = True, sheet_name = 'Rating Reference')
        ratings_sheet.to_excel(writer, index = True, sheet_name = 'Ratings Matrix')

    buffer.seek(0)
    workbook = openpyxl.load_workbook(buffer)
    sheet = workbook['Rating Reference']

    sheet['A15'] = "Explanation:"
    sheet['A16'] = "E: Excellent, V: Very Good, G: Good, F: Fair, P: Poor, -: Not Applicable"
    sheet['A17'] = "E/V: Between Excellent and Very Good, V/G: Between Very Good and Good, etc."

    buffer.seek(0)
    workbook.save(buffer)
    buffer.seek(0)
    
    return buffer

def generate_random_rankings(num_reviewers, num_proposals):
    r = np.zeros((num_reviewers, num_proposals), dtype=int)
    for i in range(num_reviewers):
        ranking = np.random.randint(1, 4, size=num_proposals)
        num_zeros = np.random.randint(0, 3)
        zero_positions = np.random.choice(num_proposals, num_zeros, replace=False)
        ranking[zero_positions] = 0
        r[i, :] = ranking
    r = r.T
    return r

def create_excel_from_dataframe(dataframes_with_sheets):
    workbook = openpyxl.Workbook()
    default_sheet = workbook.active
    default_sheet.title = 'Default Sheet'
    
    workbook.remove(default_sheet)
    
    for dataframe, sheet_name in dataframes_with_sheets:
        worksheet = workbook.create_sheet(title=sheet_name)
        
        with BytesIO() as temp_buffer:
            with pd.ExcelWriter(temp_buffer, engine='openpyxl') as writer:
                dataframe.to_excel(writer, index=True, sheet_name=sheet_name)
            
            temp_buffer.seek(0)
            temp_workbook = openpyxl.load_workbook(temp_buffer)
            temp_sheet = temp_workbook[sheet_name]
            
            for row in temp_sheet.iter_rows(values_only=True):
                worksheet.append(row)
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output

st.title('Panel Assignment')
st.sidebar.title('Inputs')

mos = st.radio("Method of Solution", ['Optimization', 'Logic'])

# number_input inputs (variable)
num_proposals = int(st.sidebar.number_input('Number of proposals', min_value = 1, step = 1, format = "%d"))
reviews_per_proposal = int(st.sidebar.number_input('Number of reviews per proposal', min_value = 1, step = 1, format = "%d"))
max_reviews_per_reviewer = int(st.sidebar.number_input('Maximum number of reviews per reviewer', min_value = 1, step = 1, format = "%d"))

# number_input inputs (fixed)
total_reviews = num_proposals * reviews_per_proposal
min_reviwers_val = int(np.ceil(total_reviews / max_reviews_per_reviewer))
max_reviewers_val = total_reviews
st.sidebar.markdown(f'Number of reviewers should be between {min_reviwers_val} and {max_reviewers_val}')
num_reviewers = int(st.sidebar.number_input('Number of reviewers', min_value = min_reviwers_val, max_value = max_reviewers_val, step = 1, format = "%d"))

if 'rankings' not in st.session_state:
    st.session_state['rankings'] = None

# generate random rankings matrix
if st.sidebar.button('Generate random rankings matrix'):
    st.session_state['rankings'] = generate_random_rankings(num_reviewers, num_proposals)
    st.session_state['rankings'] = st.session_state['rankings'].T
    st.write(pd.DataFrame(st.session_state['rankings'].T, columns=[f'Reviewer {i+1}' for i in range(num_reviewers)], index=[f'Proposal {i+1}' for i in range(num_proposals)]))

# generate rankings matrix template
if st.sidebar.button('Download rankings matrix template'):
    buffer = generate_rankings_matrix_template(num_reviewers, num_proposals)
    b64 = base64.b64encode(buffer.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="rankings_matrix_template.xlsx">Rankings matrix template</a>'
    st.sidebar.markdown(href, unsafe_allow_html=True)

# rankings matrix input
rankings_sheet = st.sidebar.file_uploader("Upload the filled rankings matrix template", type="xlsx")
if rankings_sheet is not None:
    rankings_df = pd.read_excel(rankings_sheet, engine='openpyxl', index_col = 0)
    if rankings_df.empty or (rankings_df.select_dtypes(include=[np.number]) == 0).all().all():
        st.error('The uploaded rankings matrix is empty (all zeros). Please upload a filled rankings matrix.')
    else:
        st.session_state['rankings'] = rankings_df.to_numpy()
        st.session_state['rankings'] = st.session_state['rankings'].T

# generate ratings matrix template
if st.sidebar.button('Download ratings matrix template'):
    buffer = generate_ratings_matrix_template(num_reviewers, num_proposals)
    b64 = base64.b64encode(buffer.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="ratings_matrix_template.xlsx">Download the ratings matrix template</a>'
    st.sidebar.markdown(href, unsafe_allow_html=True)

# user ratings input
user_ratings = st.sidebar.file_uploader("Upload the ratings", type="xlsx")
if user_ratings is not None:
    user_ratings_df = pd.read_excel(user_ratings, engine='openpyxl', index_col = 0)
    if user_ratings_df.empty:
        st.error('The uploaded ratings matrix is empty. Please upload a filled ratings matrix.')
    else:
        st.session_state['ratings'] = user_ratings_df

l_as_s = st.sidebar.radio("Use Lead as Scribe", ['Yes', 'No'])

if st.sidebar.button('Optimize'):
    if st.session_state['rankings'] is not None:
        # rankings = generate_random_rankings(num_reviewers, num_proposals)
        if np.any(st.session_state['rankings'] != 0):
            rankings = st.session_state['rankings']
            cols = [f'Reviewer {i+1}' for i in range(num_reviewers)]
            rows = [f'Proposal {i+1}' for i in range(num_proposals)]
            st.markdown('<h5>Rankings matrix:</h5>', unsafe_allow_html=True)
            st.write(pd.DataFrame(rankings.T, columns=cols, index=rows))

            start_time = time.time()

            min_reviews_per_reviewer = int(np.floor(total_reviews / num_reviewers))
            max_reviews_per_reviewer = int(np.ceil(total_reviews / num_reviewers))
            extra_reviews = total_reviews % num_reviewers

            num_vars = num_reviewers * num_proposals

            f = rankings.flatten().reshape(-1, 1)

            if mos == 'Optimization':

                if l_as_s == 'Yes':
                    # review assignment optimization
                    review_assignments, review_fval = review_opt()
                    if review_assignments is None or review_fval is None:
                        st.stop()
                    
                    lead_assignments = np.zeros((num_proposals, reviews_per_proposal))
                    lead_counts = np.zeros(num_reviewers)
                    proposals_assigned = np.zeros(num_proposals, dtype = bool)

                    total_preferences = np.sum(rankings, 1)
                    rankings_nan = rankings.astype(float)
                    rankings_nan[rankings_nan == 0] = np.nan
                    average_preferences = np.nanmean(rankings_nan, 1)
                    reviews_count = np.sum(review_assignments, 1)

                    data = np.column_stack((reviews_count.flatten(), average_preferences))
                    df = pd.DataFrame(data, columns=['reviews_count', 'average_preferences'])
                    sorted_df = df.sort_values(by=['reviews_count', 'average_preferences'], ascending=[True, True])
                    sorted_reviewers = sorted_df.index.values

                    round_robin_index = 0

                    for lead in range(num_proposals):
                        reviewer = sorted_reviewers[round_robin_index]
                        lead_counts[reviewer] += 1
                        round_robin_index = (round_robin_index + 1) % int(num_reviewers)

                    # lead assignment optimization
                    lead_assignments, lead_fval = lead_opt(lead_counts, review_assignments)
                    if lead_assignments is None or lead_fval is None:
                        st.stop()

                    # check for conflicts
                    check_review_conflicts(review_assignments)
                    check_lead_conflicts(lead_assignments)

                    # check for review count
                    check_review_count(review_assignments)  

                    # check for lead count
                    check_lead_count(lead_assignments.T)

                    # generate combined assignments
                    combined_assignments, fval_combined = make_combined_assignments(review_assignments, lead_assignments, rankings)

                    # calculate fairness metrics
                    fairness_prop_metric, fairness_metric, fairness_lsr_metric = calculate_fairness_metrics(review_assignments, lead_assignments, rankings, lead_counts)

                    # generate conflict reordered assignments
                    con_reordered_combined_assignments_df = generate_conflict_reordered_assignments(review_assignments, lead_assignments, rankings)

                    # generate leads reordered assignments
                    lead_reordered_combined_assignments_df = generate_leads_reordered_assignments(lead_assignments, combined_assignments)

                    # generate rating reordered assignments
                    ratings_table_df, rating_combined_assignments_with_total = generate_rating_reordered_assignments(combined_assignments, review_assignments)
                    
                    end_time = time.time()

                    # display results
                    column_names = [f'Proposal {i + 1}' for i in range(num_proposals)]
                    fairness_prop_count_df = pd.DataFrame(fairness_prop_metric.reshape(1, -1), index = ['Value'], columns = column_names)

                    column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    fairness_metric_df = pd.DataFrame(fairness_metric.reshape(1, -1), index = ['Value'], columns = column_names)

                    column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    fairness_lsr_metric_df = pd.DataFrame(fairness_lsr_metric.reshape(1, -1), index = ['Value'], columns=column_names)
                    
                    st.markdown('<h4>Results</h4>', unsafe_allow_html=True)

                    st.markdown(f'<h10>Objective function value for review assignment: {review_fval}</h10>', unsafe_allow_html=True)
                    st.markdown(f'<h10>Objective function value for combined assignment: {fval_combined}</h10>', unsafe_allow_html=True)
                    st.markdown(f'<h10>Time taken for optimization: {end_time - start_time:.2f} seconds</h10>', unsafe_allow_html=True)

                    proposal_labels = [f'Proposal {i+1}' for i in range(num_proposals)]
                    reviewer_labels = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    combined_assignments_df = pd.DataFrame(combined_assignments, columns = reviewer_labels, index = proposal_labels)
                    st.markdown('<h5>Combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # st.write(combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(combined_assignments_df)
                    st.write('')

                    st.markdown('<h5>Conflicts Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # st.write(con_reordered_combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(con_reordered_combined_assignments_df) 
                    st.write('')

                    st.markdown('<h5>Leads Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # st.write(lead_reordered_combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(lead_reordered_combined_assignments_df)
                    st.write('')

                    st.markdown('<h5>Ratings (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    st.write(ratings_table_df)
                    st.write('')

                    st.markdown('<h5>Rating Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # st.write(rating_combined_assignments_with_total.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(rating_combined_assignments_with_total)
                    st.write('')

                    st.markdown('<h5>Fairness metric for each Reviewer:</h5>', unsafe_allow_html=True)
                    # st.write(fairness_metric_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(fairness_metric_df)
                    st.write('')

                    st.markdown('<h5>Fairness metric for each LSR:</h5>', unsafe_allow_html=True)
                    # st.write(fairness_lsr_metric_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(fairness_lsr_metric_df)
                    st.write('')

                    reviews_count_df = pd.DataFrame(reviews_count.reshape(1, -1), index = ['Value'], columns = [f'Reviewer {i + 1}' for i in range(num_reviewers)])
                    st.markdown('<h5>Number of total reviews per reviewer:</h5>', unsafe_allow_html=True)
                    # st.write(reviews_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(reviews_count_df)
                    st.write('')

                    lead_counts_df = pd.DataFrame(lead_counts.reshape(1, -1), index = ['Value'], columns = [f'Reviewer {i + 1}' for i in range(num_reviewers)])
                    st.markdown('<h5>Number of leads per reviewer:</h5>', unsafe_allow_html=True)
                    # st.write(lead_counts_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(lead_counts_df)
                    st.write('')

                    proposal_count_df = pd.DataFrame(np.sum(review_assignments, 0).reshape(1, -1), index = ['Value'], columns = proposal_labels)
                    st.markdown('<h5>Number of reviews per proposal:</h5>', unsafe_allow_html=True)
                    # st.write(proposal_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(proposal_count_df)
                    st.write('')

                    lead_proposal_count_df = pd.DataFrame(np.sum(lead_assignments, 1).reshape(1, -1), index = ['Value'], columns = proposal_labels)
                    st.markdown('<h5>Number of leads per proposal:</h5>', unsafe_allow_html=True)
                    # st.write(lead_proposal_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(lead_proposal_count_df)
                    st.write('')

                    st.markdown('<h5>Fairness metric per proposal:</h5>', unsafe_allow_html=True)
                    # st.write(fairness_prop_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    st.write(fairness_prop_count_df)

                    # create output sheet and download results
                    dataframes_with_sheets = [
                        (combined_assignments_df, 'Combined assignment matrix'),
                        (con_reordered_combined_assignments_df, 'Conflicts Reordered combined assignment matrix'),
                        (lead_reordered_combined_assignments_df, 'Leads Reordered combined assignment matrix'),
                        (ratings_table_df, 'Ratings'),
                        (rating_combined_assignments_with_total, 'Rating Reordered combined assignment matrix'),
                        (fairness_metric_df, 'Fairness metric'),
                        (fairness_lsr_metric_df, 'Fairness LSR metric'),
                        (reviews_count_df, 'Number of total reviews per reviewer'),
                        (lead_counts_df, 'Number of leads per reviewer'),
                        (proposal_count_df, 'Number of reviews per proposal'),
                        (lead_proposal_count_df, 'Number of leads per proposal'),
                        (fairness_prop_count_df, 'Fairness metric per proposal')
                    ]

                    excel_buffer = create_excel_from_dataframe(dataframes_with_sheets)

                    if excel_buffer is not None:
                        # st.download_button(
                        #     label="Download Excel file",
                        #     data=excel_buffer,
                        #     file_name="output.xlsx",
                        #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        # )
                        base64_excel = base64.b64encode(excel_buffer.getvalue()).decode('utf-8')
                        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{base64_excel}" download="output.xlsx">Download Excel file</a>'
                        st.markdown(href, unsafe_allow_html=True)

                elif l_as_s == 'No':
                    # # review assignment optimization
                    # review_assignments, review_fval = review_opt()
                    # if review_assignments is None or review_fval is None:
                    #     st.stop()
                    
                    # # lead assignment optimization
                    # lead_assignments = np.zeros((num_proposals, reviews_per_proposal))
                    # lead_counts = np.zeros(num_reviewers)
                    # proposals_assigned = np.zeros(num_proposals, dtype = bool)

                    # total_preferences = np.sum(rankings, 1)
                    # rankings_nan = rankings.astype(float)
                    # rankings_nan[rankings_nan == 0] = np.nan
                    # average_preferences = np.nanmean(rankings_nan, 1)
                    # reviews_count = np.sum(review_assignments, 1)

                    # data = np.column_stack((reviews_count.flatten(), average_preferences))
                    # df = pd.DataFrame(data, columns=['reviews_count', 'average_preferences'])
                    # sorted_df = df.sort_values(by=['reviews_count', 'average_preferences'], ascending=[True, True])
                    # sorted_reviewers = sorted_df.index.values

                    # round_robin_index = 0

                    # for lead in range(num_proposals):
                    #     reviewer = sorted_reviewers[round_robin_index]
                    #     lead_counts[reviewer] += 1
                    #     round_robin_index = (round_robin_index + 1) % int(num_reviewers)

                    # lead_assignments, lead_fval = lead_opt(lead_counts, review_assignments)
                    # if lead_assignments is None or lead_fval is None:
                    #     st.stop()

                    # # scribe assignment optimization
                    # new_assignments = review_assignments.T
                    # new_assignments = new_assignments - lead_assignments
                    # new_assignments = new_assignments.T

                    # min_scribe = int(np.floor(num_proposals / num_reviewers))
                    # max_scribe = int(np.ceil(num_proposals / num_reviewers))

                    # num_vars = num_reviewers * num_proposals

                    # reviews_count = np.sum(new_assignments, 1)
                    # lead_counts = np.sum(lead_assignments, 0)
                    # scribe_counts0 = reviews_count - lead_counts

                    # new_rankings = rankings + scribe_counts0.T.reshape(-1, 1)

                    # f = new_rankings.flatten().reshape(-1, 1)

                    # scribe_assignments, scribe_fval = scribe_opt(max_scribe, min_scribe, new_assignments)
                    # if scribe_assignments is None or scribe_fval is None:
                    #     st.stop()

                    # scribe_counts = np.sum(scribe_assignments, 0)

                    # # check for conflicts
                    # check_review_conflicts(review_assignments)
                    # check_lead_conflicts(lead_assignments)
                    # check_scribe_conflicts(scribe_assignments)

                    # # check for review count
                    # check_review_count(review_assignments)

                    # # check for lead count
                    # check_lead_count(lead_assignments.T)

                    # # check for scribe count
                    # check_scribe_count(scribe_assignments.T)

                    # # generate combined assignments
                    # combined_assignments, fval_combined = make_combined_assignments_with_scribe(review_assignments, lead_assignments, scribe_assignments, rankings)
                    
                    # # calculate fairness metrics
                    # fairness_prop_metric, fairness_metric, fairness_l_metric, fairness_s_metric = calculate_fairness_metrics_with_scribe(review_assignments, lead_assignments, scribe_assignments, rankings, lead_counts, scribe_counts)

                    # # generate conflict reordered assignments
                    # con_reordered_combined_assignments_df = generate_conflict_reordered_assignments_with_scribe(review_assignments, lead_assignments, scribe_assignments, rankings)

                    # # generate leads reordered assignments
                    # lead_reordered_combined_assignments_df = generate_leads_reordered_assignments(lead_assignments, combined_assignments)

                    # # generate rating reordered assignments
                    # ratings_table_df, rating_combined_assignments_with_total = generate_rating_reordered_assignments(combined_assignments, review_assignments)

                    # end_time = time.time()

                    # # display results
                    # column_names = [f'Proposal {i + 1}' for i in range(num_proposals)]
                    # fairness_prop_count_df = pd.DataFrame(fairness_prop_metric.reshape(1, -1), index = ['Value'], columns = column_names)

                    # column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    # fairness_metric_df = pd.DataFrame(fairness_metric.reshape(1, -1), index = ['Value'], columns = column_names)

                    # column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    # fairness_l_metric_df = pd.DataFrame(fairness_l_metric.reshape(1, -1), index = ['Value'], columns=column_names)

                    # column_names = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    # fairness_s_metric_df = pd.DataFrame(fairness_s_metric.reshape(1, -1), index = ['Value'], columns=column_names)

                    # st.markdown('<h4>Results</h4>', unsafe_allow_html=True)

                    # st.markdown(f'<h10>Objective function value for review assignment: {review_fval}</h10>', unsafe_allow_html=True)
                    # st.markdown(f'<h10>Objective function value for combined assignment: {fval_combined}</h10>', unsafe_allow_html=True)
                    # st.markdown(f'<h10>Time taken for optimization: {end_time - start_time:.2f} seconds</h10>', unsafe_allow_html=True)
                    
                    # proposal_labels = [f'Proposal {i+1}' for i in range(num_proposals)]
                    # reviewer_labels = [f'Reviewer {i+1}' for i in range(num_reviewers)]
                    # combined_assignments_df = pd.DataFrame(combined_assignments, columns = reviewer_labels, index = proposal_labels)
                    # st.markdown('<h5>Combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # # st.write(combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(combined_assignments_df)
                    # st.write('')

                    # st.markdown('<h5>Conflicts Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # # st.write(con_reordered_combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(con_reordered_combined_assignments_df)
                    # st.write('')

                    # st.markdown('<h5>Leads Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # # st.write(lead_reordered_combined_assignments_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(lead_reordered_combined_assignments_df)
                    # st.write('')

                    # st.markdown('<h5>Ratings (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # st.write(ratings_table_df)
                    # st.write('')

                    # st.markdown('<h5>Rating Reordered combined assignment matrix (Proposals x Reviewers):</h5>', unsafe_allow_html=True)
                    # # st.write(rating_combined_assignments_with_total.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(rating_combined_assignments_with_total)
                    # st.write('')

                    # st.markdown('<h5>Fairness metric for each Reviewer:</h5>', unsafe_allow_html=True)
                    # # st.write(fairness_metric_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(fairness_metric_df)
                    # st.write('')

                    # st.markdown('<h5>Fairness metric for each LSR:</h5>', unsafe_allow_html=True)
                    # # st.write(fairness_l_metric_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(fairness_l_metric_df)
                    # st.write('')

                    # st.markdown('<h5>Fairness metric for each Scribe:</h5>', unsafe_allow_html=True)
                    # # st.write(fairness_s_metric_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(fairness_s_metric_df)
                    # st.write('')

                    # reviews_count_df = pd.DataFrame(reviews_count.reshape(1, -1), index = ['Value'], columns = [f'Reviewer {i + 1}' for i in range(num_reviewers)])
                    # st.markdown('<h5>Number of total reviews per reviewer:</h5>', unsafe_allow_html=True)
                    # # st.write(reviews_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(reviews_count_df)
                    # st.write('')

                    # lead_counts_df = pd.DataFrame(lead_counts.reshape(1, -1), index = ['Value'], columns = [f'Reviewer {i + 1}' for i in range(num_reviewers)])
                    # st.markdown('<h5>Number of leads per reviewer:</h5>', unsafe_allow_html=True)
                    # # st.write(lead_counts_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(lead_counts_df)
                    # st.write('')

                    # scribe_counts_df = pd.DataFrame(scribe_counts.reshape(1, -1), index = ['Value'], columns = [f'Reviewer {i + 1}' for i in range(num_reviewers)])
                    # st.markdown('<h5>Number of scribes per reviewer:</h5>', unsafe_allow_html=True)
                    # # st.write(scribe_counts_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(scribe_counts_df)

                    # proposal_count_df = pd.DataFrame(np.sum(review_assignments, 0).reshape(1, -1), index = ['Value'], columns = proposal_labels)
                    # st.markdown('<h5>Number of reviews per proposal:</h5>', unsafe_allow_html=True)
                    # # st.write(proposal_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(proposal_count_df)
                    # st.write('')

                    # lead_proposal_count_df = pd.DataFrame(np.sum(lead_assignments, 1).reshape(1, -1), index = ['Value'], columns = proposal_labels)
                    # st.markdown('<h5>Number of leads per proposal:</h5>', unsafe_allow_html=True)
                    # # st.write(lead_proposal_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(lead_proposal_count_df)
                    # st.write('')

                    # scribe_proposal_count_df = pd.DataFrame(np.sum(scribe_assignments, 1).reshape(1, -1), index = ['Value'], columns = proposal_labels)
                    # st.markdown('<h5>Number of scribes per proposal:</h5>', unsafe_allow_html=True)
                    # # st.write(scribe_proposal_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(scribe_proposal_count_df)
                    # st.write('')

                    # st.markdown('<h5>Fairness metric per proposal:</h5>', unsafe_allow_html=True)
                    # # st.write(fairness_prop_count_df.to_html(classes='full-width-table'), unsafe_allow_html=True)
                    # st.write(fairness_prop_count_df)
                    # st.write('')

                    # # create output sheet and download results
                    # dataframes_with_sheets = [
                    #     (combined_assignments_df, 'Combined assignment matrix'),
                    #     (con_reordered_combined_assignments_df, 'Conflicts Reordered combined assignment matrix'),
                    #     (lead_reordered_combined_assignments_df, 'Leads Reordered combined assignment matrix'),
                    #     (ratings_table_df, 'Ratings'),
                    #     (rating_combined_assignments_with_total, 'Rating Reordered combined assignment matrix'),
                    #     (fairness_metric_df, 'Fairness metric'),
                    #     (fairness_l_metric_df, 'Fairness L metric'),
                    #     (fairness_s_metric_df, 'Fairness S metric'),
                    #     (reviews_count_df, 'Number of total reviews per reviewer'),
                    #     (lead_counts_df, 'Number of leads per reviewer'),
                    #     (scribe_counts_df, 'Number of scribes per reviewer'),
                    #     (proposal_count_df, 'Number of reviews per proposal'),
                    #     (lead_proposal_count_df, 'Number of leads per proposal'),
                    #     (scribe_proposal_count_df, 'Number of scribes per proposal'),
                    #     (fairness_prop_count_df, 'Fairness metric per proposal')
                    # ]

                    # excel_buffer = create_excel_from_dataframe(dataframes_with_sheets)

                    # if excel_buffer is not None:
                    #     st.download_button(
                    #         label="Download Excel file",
                    #         data=excel_buffer,
                    #         file_name="output.xlsx",
                    #         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    #     )
                    pass
                
                elif mos == 'Logic':
                    pass 

        else:
            st.error('The rankings matrix is empty.')
    else:
        st.error('The rankings matrix is empty or has not been uploaded.')